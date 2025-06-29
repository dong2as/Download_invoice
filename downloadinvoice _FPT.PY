import os
import time
import openpyxl
import xml.etree.ElementTree as ET

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service

# === Cấu hình ===
DOWNLOAD_DIR = os.path.join(os.getcwd(), "downloads")
URL = "https://tracuuhoadon.fpt.com/"
CHROME_DRIVER_PATH = r"C:\Users\Admin\Documents\Visual Studio 2022\DownloadInvoice\chromedriver.exe"

# === Tạo trình duyệt ===
def tao_trinh_duyet():
    if not os.path.exists(DOWNLOAD_DIR):
        os.makedirs(DOWNLOAD_DIR)

    chrome_options = Options()
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True
    })
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-gpu")

    service = Service(executable_path=CHROME_DRIVER_PATH)
    return webdriver.Chrome(service=service, options=chrome_options)

# === Đọc dữ liệu Excel ===
def doc_file_excel(input_path):
    wb = openpyxl.load_workbook(input_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        mst = row[1]
        ma_tra_cuu = row[2]
        if mst and ma_tra_cuu:
            data.append((str(mst).strip(), str(ma_tra_cuu).strip()))
    return data

# === Tách nhỏ quá trình tra cứu ===
def openfile(driver, url):
    driver.get(url)
    wait = WebDriverWait(driver, 20)
    return wait

def nhapma(wait, ma_so_thue, ma_tra_cuu):
    input_mst = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//label[contains(text(), 'Mã số thuế')]/following-sibling::input")))
    input_mst.clear()
    input_mst.send_keys(ma_so_thue)

    input_tra_cuu = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//label[contains(text(), 'Mã tra cứu')]/following-sibling::input")))
    input_tra_cuu.clear()
    input_tra_cuu.send_keys(ma_tra_cuu)

def nuttracuu(wait):
    btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Tra cứu')]")))
    btn.click()

def download(driver):
    try:
        wait = WebDriverWait(driver, 20)
        btn_tai = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[.//span[contains(@class,'mdi-xml')]]")))

        file_truoc = set(os.listdir(DOWNLOAD_DIR))
        btn_tai.click()

        for _ in range(20):
            time.sleep(1)
            files_sau = set(os.listdir(DOWNLOAD_DIR))
            files_moi = list(files_sau - file_truoc)
            for f in files_moi:
                if f.endswith(".xml") and not f.endswith(".crdownload"):
                    return "success", os.path.join(DOWNLOAD_DIR, f)

        return "fail", None
    except Exception as e:
        print(f"❌ Lỗi tải file: {e}")
        return "fail", None

# === Hàm tổng hợp tra cứu ===
def tra_cuu_hoa_don(driver, url, ma_tra_cuu, ma_so_thue):
    try:
        wait = openfile(driver, url)
        nhapma(wait, ma_so_thue, ma_tra_cuu)
        nuttracuu(wait)
        return download(driver)
    except Exception as e:
        print(f"❌ Lỗi tra cứu ({ma_so_thue} - {ma_tra_cuu}): {e}")
        return "fail", None

# === Đọc XML ===
def doc_thong_tin_tu_xml(xml_path):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        ns = {}
        if '}' in root.tag:
            ns['ns'] = root.tag.split('}')[0].strip('{')
            prefix = 'ns:'
        else:
            prefix = ''

        def get_text(path_options):
            for path in path_options:
                el = root.find(f".//{prefix}{path}", namespaces=ns)
                if el is not None and el.text:
                    return el.text.strip()
            return ""

        return {
            'SoHoaDon': get_text(['TTChung/SHDon', 'TTChung/SoHD']),
            'DonViBanHang': get_text(['NBan/Ten']),
            'MSTBan': get_text(['NBan/MST']),
            'DiaChiBan': get_text(['NBan/DChi']),
            'SoTaiKhoan': get_text(['NBan/STKNHang', 'NBan/STK']),
            'NguoiMua': get_text(['NMua/Ten']),
            'DiaChiMua': get_text(['NMua/DChi']),
            'MSTMua': get_text(['NMua/MST'])
        }
    except Exception as e:
        print(f"❌ Lỗi đọc XML {xml_path}: {e}")
        return None

# === Ghi kết quả ra Excel ===
def ghi_ket_qua_excel(data_list, file_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    headers = [
        "Mã số thuế", "Mã tra cứu", "Trạng thái",
        "Số HĐ", "Đơn vị bán", "MST bán", "Địa chỉ bán", "STK bán",
        "Người mua", "Địa chỉ mua", "MST mua"
    ]
    sheet.append(headers)
    for row in data_list:
        sheet.append(row)
    wb.save(file_path)

# === Main ===
if __name__ == "__main__":
    driver = tao_trinh_duyet()
    danh_sach = doc_file_excel("input.xlsx")
    ket_qua = []

    for mst, ma_tra_cuu in danh_sach:
        print(f"🔎 Đang tra cứu: {mst} - {ma_tra_cuu}")
        status, xml_path = tra_cuu_hoa_don(driver, URL, ma_tra_cuu, mst)

        if status == "fail" or not xml_path:
            ket_qua.append([mst, ma_tra_cuu, "fail"] + [""] * 8)
            print("❌ Tra cứu thất bại hoặc không tải được file XML.")
        else:
            thong_tin = doc_thong_tin_tu_xml(xml_path)
            if thong_tin:
                ket_qua.append([
                    mst, ma_tra_cuu, "success",
                    thong_tin.get('SoHoaDon', ''),
                    thong_tin.get('DonViBanHang', ''),
                    thong_tin.get('MSTBan', ''),
                    thong_tin.get('DiaChiBan', ''),
                    thong_tin.get('SoTaiKhoan', ''),
                    thong_tin.get('NguoiMua', ''),
                    thong_tin.get('DiaChiMua', ''),
                    thong_tin.get('MSTMua', '')
                ])
                print("✅ Tra cứu thành công.")
            else:
                ket_qua.append([mst, ma_tra_cuu, "fail"] + [""] * 8)
                print("⚠️ Lỗi phân tích XML.")

    driver.quit()
    ghi_ket_qua_excel(ket_qua, "output.xlsx")
    print("🎉 Hoàn tất. Kết quả lưu tại: output.xlsx")
